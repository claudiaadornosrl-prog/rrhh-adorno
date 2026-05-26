"""
═══════════════════════════════════════════════════════════════════════
 09_clean_start_banco.py — Clean start del banco de minutos

 Hace 2 cosas:
   1. BORRA todos los movimientos actuales de rrhh_banco_minutos
   2. IMPORTA los saldos iniciales desde Permisos.xlsx como un único
      movimiento por empleado, fecha 2026-05-31, tipo='saldo_inicial'.

 Excel esperado: /sessions/.../uploads/Permisos.xlsx
   3 hojas: Oficina, Unicenter, Alcorta
   Columnas: B=Empleado, C=Permiso, D=Extra, E=Saldo (E=D-C)

 USO:
   python 09_clean_start_banco.py             # DRY-RUN
   python 09_clean_start_banco.py --aplicar   # Escribir
═══════════════════════════════════════════════════════════════════════
"""
import sys
if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass

import os, json, argparse
from pathlib import Path
from urllib import request as urlrequest
from urllib.parse import quote
from collections import defaultdict

try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).parent.parent / 'sync_anviz' / '.env')
except ImportError:
    pass

SUPA_URL = os.environ.get('SUPABASE_URL', 'https://kwwiykssrpabncpqtmwi.supabase.co')
SUPA_KEY = os.environ.get('SUPABASE_SERVICE_KEY')

H = {
    "apikey": SUPA_KEY or '',
    "Authorization": f"Bearer {SUPA_KEY or ''}",
    "Content-Type": "application/json; charset=utf-8",
}

FECHA_SALDO_INICIAL = '2026-05-31'

# Ruta común del Excel
EXCEL_CANDIDATOS = [
    '/sessions/stoic-dreamy-pasteur/mnt/uploads/Permisos.xlsx',
    Path(__file__).parent.parent.parent.parent.parent / 'uploads' / 'Permisos.xlsx',
    Path(__file__).parent / 'Permisos.xlsx',
]


def _encode_url(p): return quote(p, safe="/?=&.,:%-_*+")

def fetch_get(path):
    req = urlrequest.Request(f"{SUPA_URL}/rest/v1/{_encode_url(path)}", headers=H)
    with urlrequest.urlopen(req, timeout=30) as r: return json.loads(r.read())

def fetch_post(path, body):
    data = json.dumps(body, ensure_ascii=False).encode('utf-8')
    req = urlrequest.Request(f"{SUPA_URL}/rest/v1/{_encode_url(path)}",
                             data=data, headers={**H, "Prefer": "return=representation"},
                             method='POST')
    try:
        with urlrequest.urlopen(req, timeout=30) as r: return json.loads(r.read())
    except Exception as e:
        if hasattr(e, 'read'):
            print(f"   ❌ POST error: {e.read().decode('utf-8','replace')[:300]}")
        raise

def fetch_delete(path):
    req = urlrequest.Request(f"{SUPA_URL}/rest/v1/{_encode_url(path)}", headers=H, method='DELETE')
    with urlrequest.urlopen(req, timeout=30) as r: return r.read()


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--aplicar', action='store_true')
    ap.add_argument('--excel', help='Ruta al Permisos.xlsx (opcional)')
    args = ap.parse_args()

    if not SUPA_KEY: sys.exit("❌ Falta SUPABASE_SERVICE_KEY")

    # Localizar Excel
    excel_path = None
    if args.excel:
        excel_path = Path(args.excel)
    else:
        for c in EXCEL_CANDIDATOS:
            p = Path(c)
            if p.exists():
                excel_path = p; break
    if not excel_path or not excel_path.exists():
        sys.exit(f"❌ No encuentro Permisos.xlsx. Pasalo con --excel <ruta>")
    print(f"📂 Excel: {excel_path}")

    from openpyxl import load_workbook
    wb = load_workbook(excel_path, data_only=True)
    print(f"   Hojas: {wb.sheetnames}")

    # Leer saldos por hoja
    saldos_excel = []   # [{apellido, local, saldo_min}]
    HOJA_A_LOCAL = {
        'oficina': 'oficina',
        'unicenter': 'unicenter',
        'alcorta': 'alcorta',
    }
    for sh in wb.sheetnames:
        local = HOJA_A_LOCAL.get(sh.strip().lower())
        if not local:
            print(f"   ⚠ Hoja '{sh}' ignorada")
            continue
        ws = wb[sh]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 5: continue
            cell_b = row[1]
            if cell_b is None: continue
            apellido = str(cell_b).strip() if not isinstance(cell_b, str) else cell_b.strip()
            saldo = row[4]
            if not apellido or saldo is None: continue
            # Saltar header repetido o filas no de empleado
            if apellido.lower() in ('empleado', 'nombre'): continue
            try:
                saldo_int = int(float(saldo))
            except (ValueError, TypeError):
                continue
            saldos_excel.append({'apellido': apellido, 'local': local, 'saldo_min': saldo_int})

    print(f"\n   {len(saldos_excel)} saldos en el Excel:")
    for s in saldos_excel:
        print(f"     [{s['local']:10}] {s['apellido']:15} → {s['saldo_min']:+6} min")

    # Cargar empleados activos
    empleados = fetch_get("rrhh_empleados?select=id,nombre_completo,local,estado&estado=eq.activo&ficha=eq.true")
    print(f"\n📥 {len(empleados)} empleados activos\n")

    # Mapeo apellido → empleado
    def buscar(apellido, local):
        key = apellido.strip().lower()
        candidatos = [e for e in empleados if e['local'] == local]
        for e in candidatos:
            nombre_norm = (e['nombre_completo'] or '').lower().replace(',', '')
            if key in nombre_norm.split():
                return e
            # Buscar primer token del apellido en el nombre
            partes = nombre_norm.split()
            if partes and partes[0] == key:
                return e
        # Match parcial
        for e in candidatos:
            if key in (e['nombre_completo'] or '').lower():
                return e
        return None

    # Planificar inserts
    inserts = []
    no_match = []
    for s in saldos_excel:
        emp = buscar(s['apellido'], s['local'])
        if not emp:
            no_match.append(s)
            continue
        inserts.append({
            'empleado_id': emp['id'],
            'fecha': FECHA_SALDO_INICIAL,
            'minutos': s['saldo_min'],
            'tipo': 'saldo_inicial',
            'referencia_tipo': 'permisos_xlsx',
            'observaciones': f'Saldo inicial importado desde Permisos.xlsx ({s["apellido"]})',
            'creado_por': 'import-clean-start',
        })

    print(f"━━━ Plan ━━━")
    print(f"   Movimientos a insertar: {len(inserts)}")
    for i in inserts:
        nom = next((e['nombre_completo'] for e in empleados if e['id'] == i['empleado_id']), '?')
        print(f"     [{i['fecha']}] {nom[:32]:32}  {i['minutos']:+6} min")
    if no_match:
        print(f"\n   ⚠ Sin match en DB ({len(no_match)}):")
        for s in no_match:
            print(f"     {s['apellido']} ({s['local']}) saldo={s['saldo_min']}")

    print(f"\n   Acciones:")
    print(f"     1. DELETE FROM rrhh_banco_minutos  (borra TODOS los movimientos)")
    print(f"     2. INSERT {len(inserts)} movimientos saldo_inicial fecha {FECHA_SALDO_INICIAL}")

    if not args.aplicar:
        print(f"\n🔵 DRY-RUN — no se escribe. Para aplicar:")
        print(f"   python {Path(__file__).name} --aplicar")
        return

    # CONFIRMACIÓN EXPLÍCITA — esto borra todo el banco
    print(f"\n⚠ Se va a BORRAR todo rrhh_banco_minutos y reimportar saldos iniciales.")
    confirma = input("   Escribí 'CLEAN START' para confirmar: ").strip()
    if confirma != 'CLEAN START':
        print("Cancelado.")
        return

    # Borrar todo
    print(f"\n🗑 Borrando rrhh_banco_minutos...")
    # PostgREST necesita un filtro para DELETE — usamos id>=0 que matchea todo
    fetch_delete("rrhh_banco_minutos?id=gte.0")
    print(f"   ✓ Borrado")

    # Insertar saldos iniciales en lote
    print(f"\n📤 Insertando saldos iniciales...")
    if inserts:
        fetch_post("rrhh_banco_minutos", inserts)
    print(f"   ✓ {len(inserts)} movimientos creados")

    print(f"\n✅ Clean start completo. El banco arranca con los saldos del Excel.")
    print(f"   📌 De ahora en más, solo procesarMes a partir de junio 2026 va a")
    print(f"      modificar el banco (eso requiere el cambio en index.html — Claude lo aplica).")


if __name__ == '__main__':
    main()
