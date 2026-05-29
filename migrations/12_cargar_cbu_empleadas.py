r"""
═══════════════════════════════════════════════════════════════════════
 12_cargar_cbu_empleadas.py
 Lee un Excel de "Detalle de registros" de Galicia Office (el archivo
 que el portal te devuelve después de hacer las acreditaciones) y
 carga los números de cuenta de Galicia al campo `cbu` de cada
 empleada en rrhh_empleados.

 Formato esperado del Excel:
   - Fila 0: encabezado "Detalle de registros para la operacion ..."
   - Fila 1: headers (Nombre, NroCuenta | CBU, Fecha, Importe, Estado, ...)
   - Fila 2+: una empleada por fila

 USO:
   python 12_cargar_cbu_empleadas.py <archivo.xlsx>            # DRY-RUN
   python 12_cargar_cbu_empleadas.py <archivo.xlsx> --aplicar  # Actualiza
═══════════════════════════════════════════════════════════════════════
"""
import sys
if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass

import os, json, re, argparse
from pathlib import Path
from urllib import request as urlrequest
from urllib.parse import quote

try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).parent.parent / 'sync_anviz' / '.env')
except ImportError:
    pass

try:
    import openpyxl
except ImportError:
    sys.exit("[ERROR] Falta openpyxl. Instalalo con: pip install openpyxl")

SUPA_URL = os.environ.get('SUPABASE_URL', 'https://kwwiykssrpabncpqtmwi.supabase.co')
SUPA_KEY = os.environ.get('SUPABASE_SERVICE_KEY')

H = {
    "apikey": SUPA_KEY or '',
    "Authorization": f"Bearer {SUPA_KEY or ''}",
    "Content-Type": "application/json; charset=utf-8",
}


def fetch_get(path):
    req = urlrequest.Request(f"{SUPA_URL}/rest/v1/{quote(path, safe='/?=&.,:%-_*+')}", headers=H)
    with urlrequest.urlopen(req, timeout=30) as r:
        return json.loads(r.read())

def fetch_patch(path, body):
    data = json.dumps(body, ensure_ascii=False).encode('utf-8')
    req = urlrequest.Request(
        f"{SUPA_URL}/rest/v1/{quote(path, safe='/?=&.,:%-_*+')}",
        data=data,
        headers={**H, "Prefer": "return=minimal"},
        method='PATCH',
    )
    with urlrequest.urlopen(req, timeout=30): return True


def normalizar(s):
    """Normaliza para matching: minúsculas, sin acentos, sin comas, tokens limpios."""
    if not s: return set()
    s = str(s).lower().strip()
    # Quitar acentos
    s = (s.replace('á','a').replace('é','e').replace('í','i').replace('ó','o').replace('ú','u')
           .replace('ñ','n').replace(',', ' '))
    # Solo letras y espacios
    s = re.sub(r'[^a-z\s]+', ' ', s)
    return set(t for t in s.split() if len(t) >= 3)


def matchear(nombre_excel, empleados):
    """Busca el empleado que mejor matchea por intersección de tokens."""
    tokens_excel = normalizar(nombre_excel)
    if not tokens_excel: return None
    mejor, mejor_score = None, 0
    for e in empleados:
        tokens_emp = normalizar(e['nombre_completo'])
        # Score = cuántos tokens del nombre Excel están en el padrón
        score = len(tokens_excel & tokens_emp)
        if score > mejor_score:
            mejor_score = score
            mejor = e
    # Necesitamos al menos 2 tokens en común para considerar match
    return mejor if mejor_score >= 2 else None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('archivo', help='Path al .xlsx de Galicia Office')
    ap.add_argument('--aplicar', action='store_true', help='Aplica los cambios (sino dry-run)')
    args = ap.parse_args()

    if not SUPA_KEY:
        sys.exit("[ERROR] Falta SUPABASE_SERVICE_KEY en .env")

    archivo = Path(args.archivo)
    if not archivo.exists():
        sys.exit(f"[ERROR] No existe el archivo: {archivo}")

    print(f"Archivo: {archivo}")
    wb = openpyxl.load_workbook(archivo, data_only=True)
    ws = wb.active
    print(f"Hoja: {ws.title} ({ws.max_row} filas)")

    # Empleados activos
    print(f"Cargando empleados del padrón...")
    empleados = fetch_get("rrhh_empleados?select=id,nombre_completo,apellido,nombre,cbu&estado=eq.activo")
    print(f"   {len(empleados)} empleados activos")

    # Localizar fila de headers (la que contiene "Nombre" y "NroCuenta")
    fila_header = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        cells = [str(c).lower() if c else '' for c in row]
        if any('nombre' in c for c in cells) and any('cuenta' in c or 'cbu' in c for c in cells):
            fila_header = i
            headers = cells
            break
    if fila_header is None:
        sys.exit("[ERROR] No se encontró fila de headers con 'Nombre' y 'NroCuenta/CBU'")
    print(f"Header en fila {fila_header}: {headers}")

    # Detectar índices de las columnas relevantes
    idx_nombre  = next((i for i, c in enumerate(headers) if 'nombre' in c and 'cuenta' not in c), None)
    idx_cuenta  = next((i for i, c in enumerate(headers) if 'cuenta' in c or 'cbu' in c), None)
    if idx_nombre is None or idx_cuenta is None:
        sys.exit(f"[ERROR] No se encontraron columnas. nombre={idx_nombre}, cuenta={idx_cuenta}")

    # Procesar filas de datos
    matches, ambiguos, sin_match, ya_cargados, actualizados = [], [], [], [], []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= fila_header: continue
        nombre_excel = row[idx_nombre]
        cuenta = row[idx_cuenta]
        if not nombre_excel or not cuenta: continue
        # Limpiar la cuenta: solo dígitos
        cuenta_str = re.sub(r'[^0-9]', '', str(cuenta))
        if not cuenta_str: continue

        emp = matchear(nombre_excel, empleados)
        if not emp:
            sin_match.append((nombre_excel, cuenta_str))
            continue

        if emp.get('cbu') == cuenta_str:
            ya_cargados.append((emp['nombre_completo'], cuenta_str))
            continue

        matches.append((emp, cuenta_str, nombre_excel))

    print(f"\n========== MATCHES ==========")
    for emp, cuenta, nombre_orig in matches:
        viejo = emp.get('cbu') or '(vacío)'
        flag = '↻' if emp.get('cbu') else '+'
        print(f"  {flag} {emp['nombre_completo']:40s}  {viejo:>15s}  →  {cuenta}")

    if ya_cargados:
        print(f"\n========== YA CARGADOS (sin cambios) ==========")
        for nombre, cuenta in ya_cargados:
            print(f"  ✓ {nombre:40s}  {cuenta}")

    if sin_match:
        print(f"\n========== SIN MATCH ==========")
        for nombre, cuenta in sin_match:
            print(f"  ✗ {nombre}  ({cuenta})")

    print(f"\nResumen:")
    print(f"  A actualizar:    {len(matches)}")
    print(f"  Ya cargados:     {len(ya_cargados)}")
    print(f"  Sin match:       {len(sin_match)}")

    if not args.aplicar:
        print(f"\nDRY-RUN. Si todo se ve bien, corré con --aplicar para guardar en la base.")
        return

    if not matches:
        print(f"\nNada para actualizar.")
        return

    print(f"\nActualizando en Supabase...")
    for emp, cuenta, _ in matches:
        try:
            fetch_patch(f"rrhh_empleados?id=eq.{emp['id']}", {'cbu': cuenta})
            actualizados.append(emp['nombre_completo'])
            print(f"  OK  {emp['nombre_completo']}")
        except Exception as e:
            print(f"  ERROR  {emp['nombre_completo']}: {e}")

    print(f"\n{len(actualizados)} empleados actualizados.")


if __name__ == '__main__':
    main()
