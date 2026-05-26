"""
═══════════════════════════════════════════════════════════════════════
 05_importar_vacaciones.py — Importar vacaciones desde Excel a Supabase

 USO:
   python 05_importar_vacaciones.py <archivo.xlsx> [--dry-run]

 FORMATO ESPERADO DEL EXCEL:
   El script intenta detectar el formato automáticamente. Soporta:

   Formato A (recomendado) — una fila por período de vacaciones:
     | Empleado          | DNI        | Desde       | Hasta       | Observ.        |
     | NOGUERA ADRIAN    | 30123456   | 2026-01-05  | 2026-01-20  | Verano         |
     | BENITEZ ROMINA    | 28987654   | 2026-05-19  | 2026-05-22  | -              |

   Variantes aceptadas en nombres de columnas:
     - Empleado: "Empleado", "Nombre", "Vendedora", "Empleada"
     - DNI: "DNI", "Dni", "Documento"   (opcional si el nombre matchea)
     - Desde: "Desde", "Fecha desde", "Inicio", "Desde fecha"
     - Hasta: "Hasta", "Fecha hasta", "Fin", "Hasta fecha"
     - Observaciones: "Obs", "Observ", "Observaciones", "Motivo", "Detalle"

 EFECTOS:
   - Crea filas en rrhh_vacaciones_movimientos con estado='tomada'
   - Actualiza dias_tomados en rrhh_vacaciones por (empleado, año)
   - NO duplica: si ya existe un movimiento (empleado, desde, hasta), lo salta
═══════════════════════════════════════════════════════════════════════
"""
import sys
if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass

import os, json, argparse
from pathlib import Path
from urllib import request as urlrequest
from datetime import date, datetime
from collections import defaultdict

try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).parent.parent / 'sync_anviz' / '.env')
except ImportError:
    pass

SUPA_URL = os.environ.get('SUPABASE_URL', 'https://kwwiykssrpabncpqtmwi.supabase.co')
SUPA_KEY = os.environ.get('SUPABASE_SERVICE_KEY')

H = {
    "apikey": SUPA_KEY or '',
    "Authorization": f"Bearer {SUPA_KEY or ''}",
    "Content-Type": "application/json",
}

# Variantes aceptadas de nombres de columna
ALIAS = {
    'empleado': ['empleado', 'nombre', 'vendedora', 'empleada', 'nombre completo'],
    'dni':      ['dni', 'documento', 'doc'],
    'desde':    ['desde', 'fecha desde', 'inicio', 'desde fecha', 'fecha inicio', 'fecha_desde'],
    'hasta':    ['hasta', 'fecha hasta', 'fin', 'hasta fecha', 'fecha fin', 'fecha_hasta'],
    'obs':      ['obs', 'observ', 'observaciones', 'motivo', 'detalle', 'comentario'],
}


def normalizar(s):
    if s is None: return ''
    return str(s).strip().lower().replace('  ', ' ')


def detectar_columnas(headers):
    """Mapea las columnas del Excel a nuestras claves canónicas."""
    out = {}
    for i, h in enumerate(headers):
        hn = normalizar(h)
        for clave, aliases in ALIAS.items():
            if hn in aliases:
                out[clave] = i
                break
    return out


def fecha_to_iso(v):
    if v is None or v == '': return None
    if isinstance(v, (date, datetime)):
        return v.strftime('%Y-%m-%d')
    s = str(v).strip()
    # Intentar varios formatos
    for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d', '%d/%m/%y', '%d-%m-%y']:
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return None


def fetch_get(path):
    req = urlrequest.Request(f"{SUPA_URL}/rest/v1/{path}", headers=H)
    with urlrequest.urlopen(req, timeout=30) as r:
        return json.loads(r.read())


def fetch_post(path, body):
    data = json.dumps(body).encode('utf-8')
    req = urlrequest.Request(f"{SUPA_URL}/rest/v1/{path}",
                             data=data, headers={**H, "Prefer": "return=representation"},
                             method='POST')
    with urlrequest.urlopen(req, timeout=30) as r:
        return json.loads(r.read())


def fetch_patch(path, body):
    data = json.dumps(body).encode('utf-8')
    req = urlrequest.Request(f"{SUPA_URL}/rest/v1/{path}",
                             data=data, headers={**H, "Prefer": "return=representation"},
                             method='PATCH')
    with urlrequest.urlopen(req, timeout=30) as r:
        return json.loads(r.read())


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('archivo', help='Ruta al .xlsx con vacaciones')
    ap.add_argument('--dry-run', action='store_true', help='No escribir, solo simular')
    ap.add_argument('--sheet', help='Nombre de hoja (opcional)')
    args = ap.parse_args()

    if not SUPA_KEY:
        sys.exit("❌ Falta SUPABASE_SERVICE_KEY en sync_anviz/.env")

    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("❌ Falta openpyxl. Corré: pip install openpyxl")

    path = Path(args.archivo)
    if not path.exists():
        sys.exit(f"❌ No existe el archivo {path}")

    print(f"📂 Leyendo {path}")
    wb = load_workbook(path, data_only=True)
    sheet_name = args.sheet or wb.sheetnames[0]
    print(f"📄 Hoja: {sheet_name}")
    ws = wb[sheet_name]

    # Leer headers de la primera fila no-vacía
    rows = list(ws.iter_rows(values_only=True))
    header_row_idx = 0
    for i, r in enumerate(rows):
        if any(c not in (None, '') for c in r):
            header_row_idx = i
            break
    headers = list(rows[header_row_idx])
    cols = detectar_columnas(headers)

    print(f"\nColumnas detectadas:")
    for k, idx in cols.items():
        print(f"   {k:10} → columna {idx} ({headers[idx]!r})")

    if 'desde' not in cols or 'hasta' not in cols:
        sys.exit("❌ No se detectaron las columnas 'Desde' y 'Hasta'. Revisá los headers del Excel.")
    if 'empleado' not in cols and 'dni' not in cols:
        sys.exit("❌ Necesito al menos columna 'Empleado' o 'DNI' para identificar a la persona.")

    # Cargar empleados de la DB
    print(f"\n📥 Cargando empleados de la DB...")
    empleados = fetch_get("rrhh_empleados?select=id,nombre_completo,dni,estado&estado=eq.activo")
    print(f"   {len(empleados)} empleados activos")

    by_dni = { (e.get('dni') or '').strip(): e for e in empleados if e.get('dni') }
    by_nombre = { normalizar(e.get('nombre_completo')): e for e in empleados }
    # Por primer apellido
    by_apellido = defaultdict(list)
    for e in empleados:
        n = normalizar(e.get('nombre_completo'))
        partes = n.replace(',', '').split()
        if partes:
            by_apellido[partes[0]].append(e)

    def buscar_empleado(emp_text, dni_text):
        if dni_text:
            d = str(dni_text).strip().replace('.', '').replace('-', '')
            if d in by_dni:
                return by_dni[d]
        if emp_text:
            n = normalizar(emp_text)
            if n in by_nombre:
                return by_nombre[n]
            # Buscar por primer apellido + algún token coincidente
            partes = n.replace(',', '').split()
            for p in partes:
                if p in by_apellido and len(by_apellido[p]) == 1:
                    return by_apellido[p][0]
            # Buscar por inclusión parcial
            for e in empleados:
                nb = normalizar(e.get('nombre_completo'))
                if all(t in nb for t in partes if len(t) > 2):
                    return e
        return None

    # Cargar movimientos existentes para no duplicar
    print(f"📥 Cargando movimientos existentes...")
    existentes = fetch_get("rrhh_vacaciones_movimientos?select=empleado_id,fecha_desde,fecha_hasta")
    set_existentes = {(m['empleado_id'], m['fecha_desde'], m['fecha_hasta']) for m in existentes}
    print(f"   {len(set_existentes)} movimientos ya en DB")

    # Procesar filas de datos
    nuevos = []
    saltados = 0
    no_encontrados = []
    fila_num = 0

    for r in rows[header_row_idx+1:]:
        fila_num += 1
        if not any(c not in (None, '') for c in r): continue  # fila vacía

        emp_text = r[cols['empleado']] if 'empleado' in cols else None
        dni_text = r[cols['dni']] if 'dni' in cols else None
        desde = fecha_to_iso(r[cols['desde']])
        hasta = fecha_to_iso(r[cols['hasta']])
        obs = r[cols['obs']] if 'obs' in cols else None

        if not desde or not hasta:
            print(f"   ⚠ Fila {fila_num+header_row_idx+1}: fechas inválidas → emp={emp_text}, desde={r[cols['desde']]}, hasta={r[cols['hasta']]}")
            continue

        emp = buscar_empleado(emp_text, dni_text)
        if not emp:
            no_encontrados.append((fila_num, emp_text, dni_text))
            continue

        clave = (emp['id'], desde, hasta)
        if clave in set_existentes:
            saltados += 1
            continue

        d1 = datetime.strptime(desde, '%Y-%m-%d')
        d2 = datetime.strptime(hasta, '%Y-%m-%d')
        dias = (d2 - d1).days + 1

        nuevos.append({
            'empleado_id': emp['id'],
            'fecha_desde': desde,
            'fecha_hasta': hasta,
            'dias_corridos': dias,
            'año': d1.year,
            'estado': 'tomada',
            'observaciones': str(obs) if obs else None,
            'solicitado_por': 'import-script',
            'revisado_por':   'import-script',
            'revisado_at':    datetime.utcnow().isoformat() + 'Z',
        })
        set_existentes.add(clave)

    print(f"\n📊 Resumen:")
    print(f"   - Nuevos a insertar:  {len(nuevos)}")
    print(f"   - Ya existían:        {saltados}")
    print(f"   - Empleados no enc.:  {len(no_encontrados)}")

    if no_encontrados:
        print(f"\n⚠ Empleados no encontrados (revisá nombre o DNI):")
        for f, n, d in no_encontrados:
            print(f"   - Fila {f}: nombre={n!r} dni={d!r}")

    if args.dry_run:
        print(f"\n🔵 DRY-RUN — no se escribe. Ejemplos:")
        for i, n in enumerate(nuevos[:5]):
            print(f"   {i+1}. emp={n['empleado_id']} {n['fecha_desde']}→{n['fecha_hasta']} ({n['dias_corridos']}d)")
        return

    if not nuevos:
        print("\nSin novedades para insertar.")
        return

    # Insertar en lote
    print(f"\n📤 Insertando {len(nuevos)} movimientos...")
    BATCH = 100
    insertados = 0
    for i in range(0, len(nuevos), BATCH):
        chunk = nuevos[i:i+BATCH]
        fetch_post("rrhh_vacaciones_movimientos", chunk)
        insertados += len(chunk)
        print(f"   ✓ {insertados}/{len(nuevos)}")

    # Actualizar saldos (dias_tomados) por empleado/año
    por_emp_año = defaultdict(int)
    for n in nuevos:
        por_emp_año[(n['empleado_id'], n['año'])] += n['dias_corridos']

    print(f"\n📤 Actualizando saldos en rrhh_vacaciones...")
    for (emp_id, año), dias in por_emp_año.items():
        # Buscar existente
        existing = fetch_get(f"rrhh_vacaciones?empleado_id=eq.{emp_id}&año=eq.{año}&select=id,dias_tomados")
        if existing:
            fetch_patch(f"rrhh_vacaciones?id=eq.{existing[0]['id']}", {
                'dias_tomados': (existing[0].get('dias_tomados') or 0) + dias
            })
        else:
            fetch_post("rrhh_vacaciones", [{
                'empleado_id': emp_id, 'año': año,
                'dias_correspondientes': 0,  # admin ajustará después
                'dias_tomados': dias,
            }])
        print(f"   ✓ emp={emp_id} año={año}: +{dias}d")

    print(f"\n✅ Importación completa: {insertados} movimientos cargados.")


if __name__ == '__main__':
    main()
